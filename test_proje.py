from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager . chrome import ChromeDriverManager
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec 
from time import sleep
from constants import globalConstants as gC
import openpyxl
import pytest
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
import time
import json
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.webdriver.support.ui import Select




class Test_project_pairtwo:
    def setup_method(self): 
        self.driver = webdriver.Chrome()
        self.driver.get(gC.BASE_URL)
        self.driver.maximize_window()
    def getData():
        excel= openpyxl.load_workbook("data/giris.xlsx")
        sheet= excel["successfulLogin"]
        rows=sheet.max_row
        data=[]
        for i in range(2,rows+1):
            eposta=sheet.cell(i,1).value
            şifre= sheet.cell(i,2).value
            university=sheet.cell(i,3).value
            department=sheet.cell(i,4).value
            data.append((eposta,şifre,university,department))
        return data
   
    @pytest.mark.parametrize("eposta,şifre,university,department", getData())
    def test_story(self,eposta,şifre,university,department):
#takvim
        calendar=self.driver.find_element(By.CSS_SELECTOR, gC.CALENDAR_CSS)
        calendar.click()
        calendarText=self.driver.find_element(By.XPATH,gC.CALENDAR_TEXT_XPATH)
        sleep(3)
        calendarText.is_displayed
        searchEducationText= WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.XPATH, gC.SEARCH_EDUCATION_XPATH)))
        searchEducationText.send_keys(gC.SEARCH_EDUCATION_TEXT)
        qualityAndTest=WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.XPATH, gC.QUALITY_TEST_XPATH)))
        sleep(3)
        allText=qualityAndTest.text
        searchtext=gC.SEARCH_EDUCATION_TEXT
        if searchtext in allText:
           print(f"'{searchtext}' kelimesi metinde bulunuyor.")
        else:
            print(f"'{searchtext}' kelimesi metinde bulunmuyor.")
        
        month=WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.XPATH,gC.MONTH_XPATH)))
        month.click()
        sleep(3)
        week=WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.XPATH,gC.WEEK_XPATH)))
        week.click()
        sleep(3)
        day=WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.XPATH,gC.DAY_XPATH)))
        day.click()
        sleep(3)

        closeButton = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.CSS_SELECTOR,".btn-close-white")))
        closeButton.click()
        sleep(3)
#chatbot
        iframe_element = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.XPATH,gC.ANASAYFA_IFRAME)))
        WebDriverWait(self.driver,10).until(ec.frame_to_be_available_and_switch_to_it(iframe_element)) 
        openChatBot = self.driver.find_element(By.XPATH, gC.CHATBOT_IFRAME)
        openChatBot.click()
        sleep(5)
        self.driver.switch_to.default_content()
        second_frame =WebDriverWait(self.driver,30).until(ec.visibility_of_element_located((By.XPATH, gC.CHATBOT_SECOND_IFRAME)))
        WebDriverWait(self.driver,15).until(ec.frame_to_be_available_and_switch_to_it(second_frame)) 
        sleep(6)
        self.driver.find_element(By.NAME, "message").click()
        sleep(5)
        self.driver.find_element(By.NAME, "message").send_keys("yasin Turgut")
        sleep(5)
        self.driver.find_element(By.XPATH, "//*[@id='exw-conversation-frame-body']/div/div/div/div[3]/form/div/button").click()
        sleep(10)
        self.driver.find_element(By.CSS_SELECTOR, ".exw-reply:nth-child(1)").click()
        sleep(10)
        self.driver.find_element(By.CSS_SELECTOR, ".exw-reply:nth-child(1)").click()
        sleep(5)
        self.driver.find_element(By.CSS_SELECTOR, ".exw-reply:nth-child(1)").click()
        sleep(10)
        self.driver.find_element(By.CSS_SELECTOR, ".exw-reply:nth-child(2)").click()
        sleep(6)
        self.driver.find_element(By.CSS_SELECTOR, "#emoji5 > #color > circle").click()
        sleep(5)
        self.driver.find_element(By.ID, "surveyBtn").click()
        self.driver.switch_to.default_content()
        sleep(5)

#login 
        loginButton = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.XPATH,gC.LOGINBUTTON)))
        loginButton.click()
        sleep(2)
        epostainput = WebDriverWait(self.driver,15).until(ec.visibility_of_element_located((By.XPATH, gC.name)))
        epostainput.send_keys(eposta)
        sleep(3)
        password = WebDriverWait(self.driver,15).until(ec.visibility_of_element_located((By.NAME, gC.password)))
        password.send_keys(şifre)
        loginbutton = self.driver.find_element(By.XPATH, gC.LOGGİNBUTTON_XPATH)
        loginbutton.click()
        sleep(4)
        mesaj = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.CSS_SELECTOR,"div[class='toast-body']")))
        assert mesaj.text=="• Giriş başarılı."
        sleep(3)

#eğitim-senkronders
        self.driver.execute_script("window.scrollTo(0,400)")
        sleep(2)
        egitimlerim = self.driver.find_element(By.XPATH,gC.LESSONS)
        egitimlerim.click()
        sleep(2)
        self.driver.execute_script("window.scrollTo(400,600)")
        sleep(5)
        show_more_button = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.XPATH, gC.SHOWMORE_BUTTON)))
        show_more_button.click()
        sleep(3)
        self.driver.execute_script("window.scrollTo(0,2000)")
        sleep(3)
        egitime_git = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.XPATH,gC.PROCEED_TO_TRAINING)))
        egitime_git.click()
        tamamladın = WebDriverWait(self.driver,15).until(ec.visibility_of_element_located((By.CSS_SELECTOR, gC.YOU_DID_IT)))
        tamamladın.click()
        egitimi_nasıl_tamamlayabilirim = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.XPATH,gC.HOW_CAN_I_COMPLETE_TRANING)))
        assert egitimi_nasıl_tamamlayabilirim.text == "Eğitimi nasıl tamamlayabilirim?"
        egitimi_nasıl_basarabilirim = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.XPATH,gC.HOW_CAN_I_ACHIEVE_TRANING)))
        assert egitimi_nasıl_basarabilirim.text == "Eğitimi nasıl başarabilirim?"
        geri_gel = WebDriverWait(self.driver,15).until(ec.visibility_of_element_located((By.CSS_SELECTOR, "a path")))
        geri_gel.click()
        sleep(3)
#KİŞİSEL BİLGİLERİM
        
#profil/eğitimekleme
        profilebutton =self.driver.find_element(By.XPATH, gC.PROFILE_XPATH)
        profilebutton.click()
        sleep(3)
        informationprofile = self.driver.find_element(By.XPATH, gC.INFORMATİON_PROFILE)
        informationprofile.click()
        sleep(3)
        myEducational_life =WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,gC.MYEDUCATIONALİFE)))
        myEducational_life.click()
        sleep(3)
        educational_status = self.driver.find_element(By.XPATH, gC.EDUCATIONAL_STATUS)
        educational_status.click()
        select_educational_status = self.driver.find_element(By.XPATH, gC.SELECT_EDUCATIONAL_STATUS)
        select_educational_status.click()
        sleep(2)
        university_field = self.driver.find_element(By.XPATH, gC. UNIVERSITY_FIELD)
        university_field.click()
        university_field.send_keys("Marmara Üniversitesi")
        sleep(2)
        department_title_yazılım = self.driver.find_element(By.XPATH, "//input[@placeholder='Yazılım']")
        department_title_yazılım.click()
        department_title_yazılım.send_keys("Felsefe ve Din Bilimleri")
        sleep(2)
        start_year_default = self.driver.find_element(By.XPATH,"//input[@placeholder='Başlangıç Yılınızı Seçiniz']")
        start_year_default.click()
        self.driver.find_element(By.CSS_SELECTOR, gC.BACK).click()
        self.driver.find_element(By.CSS_SELECTOR, gC. START_2015).click()
        sleep(2)
        graduation_year = self.driver.find_element(By.XPATH,gC.GRADUATİON)
        graduation_year.click()
        sleep(3)
        self.driver.find_element(By.CSS_SELECTOR, gC.MEZUNIYET_GERİGELME).click()
        self.driver.find_element(By.CSS_SELECTOR, gC.MEZUNIYETYILI_SECME).click()
        sleep(2)
        saveButton = self.driver.find_element(By.XPATH,gC.SAVEBUTTON)
        saveButton.click()
        toast_message = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.CSS_SELECTOR, ".toast-body")))
        assert toast_message.text == "• Eğitim bilgisi eklendi."
        sleep(2)

#YABANCI DİL
        
        
        
    

    

